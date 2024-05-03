from django.shortcuts import render,redirect
from .models import Utilisateur
from datetime import *
from django.db.models import *
from supervisor.modeles import *
from supervisor.modeles import SupervisorUtilisateurUserPermissions
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from .tokens import generateToken
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.utils.encoding import force_bytes, force_text
from django.contrib.auth.hashers import make_password
from django.contrib.auth.decorators import login_required
from .access import *
import csv
import os
from django.http import HttpResponse
from openpyxl import Workbook
import pandas as pd
from prophet import Prophet
from functools import wraps

# Create your views here.
@login_required(login_url='login')
def home_supervisor(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    return render(request,'supervisor/home.html')

@login_required
def ajout_vendeur(request):
     # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    if request.method == "POST":
        username = request.POST['username']
        firstname = request.POST['firstname']
        lastname = request.POST['lastname']
        email = request.POST['Email']
        role = request.POST['role']
        adr = request.POST['adr']
        sex = request.POST['sex']
        tel = request.POST['tel']
        password = request.POST['mot']
        confirmpwd = request.POST['cmot']

        if SupervisorUtilisateur.objects.filter(username=username):
            messages.add_message(request,messages.ERROR, 'Ce username est deja pris.')
            return render(request,'supervisor/ajout_personnel.html',{'messages':messages.get_messages(request)})
        
        if SupervisorUtilisateur.objects.filter(email=email):
            messages.add_message(request,messages.ERROR, 'Ce email à deja un compte.')
            return render(request,'supervisor/ajout_personnel.html',{'messages':messages.get_messages(request)})
        if len(username)<5:
            messages.add_message(request,messages.ERROR, 'Entrez un username plus long.')
            return render(request,'supervisor/ajout_personnel.html',{'messages':messages.get_messages(request)})
        if not username.isalnum():
            messages.add_message(request,messages.ERROR, 'le username doit conetnir des chiffres')
            return render(request,'supervisor/ajout_personnel.html',{'messages':messages.get_messages(request)})

        if password != confirmpwd:
            messages.add_message(request,messages.ERROR, 'Mot de pass incorrect ')
            return render(request,'supervisor/ajout_personnel.html',{'messages':messages.get_messages(request)})
        hash_password = make_password(password)
        my_user = SupervisorUtilisateur.objects.create(username=username, email=email, password=hash_password)
        my_user.first_name =firstname
        my_user.last_name = lastname
        my_user.role = role
        my_user.sex = sex
        my_user.localites = adr
        my_user.tel = tel
        my_user.is_active = True 
        my_user.save()
        messages.add_message(request,messages.SUCCESS, 'felicitation ce compte.'+ my_user.role + 'à bien été créer')
# send email when account has been created successfully
        subject = "Bienvenue à Camso corps"
        message = "Bienvenue "+ my_user.first_name + " " + my_user.last_name +"\n votre username est le suivant : "+ username + "\n votre mot de pass est le suivant :"+ password + '\n vous recevrez un autre mail de confrimation du compte veuillez gardez ces information pour vous'
        
        from_email = settings.EMAIL_HOST_USER
        to_list = [my_user.email]
        send_mail(subject, message, from_email, to_list, fail_silently=False)

# send the the confirmation email
        current_site = get_current_site(request) 
        email_suject = "Activez votre compte"
        messageConfirm = render_to_string("supervisor/email.html", {
            'name': my_user.first_name,
            'domain':current_site.domain,
            'uid': urlsafe_base64_encode(force_bytes(my_user.pk)),
            'token': generateToken.make_token(my_user)
        })       

        email = EmailMessage(
            email_suject,
            messageConfirm,
            settings.EMAIL_HOST_USER,
            [my_user.email]
        )

        email.fail_silently = False
        email.send()
    return render(request,'supervisor/ajout_personnel.html')



@login_required
def historique(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    vente = ""
    Date = ""
    taille=0
    if request.method == 'POST':
        Date = request.POST['date']
        date = datetime.strptime(Date, "%Y-%m-%d").date()
        vente = VentesAnalyse.objects.filter( date__year=date.year,date__month=date.month,date__day=date.day)
        taille = len(vente)
    else:
        today = datetime.now()
        vente = VentesAnalyse.objects.filter( date__year=today.year,date__month=today.month,date__day=today.day)
        taille = len(vente)
    return render(request, 'supervisor/ventes.html', {'ventes':vente, 'date':Date, 'taille':taille})



def ajout_gestionnaire(request):
     # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    return render(request,'supervisor/ajout_personnel.html')

@login_required
def mot(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    return render(request,'supervisor/motpass.html')

@login_required
def oublie(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    return render(request, 'supervisor/oublie.html')

@login_required
def signout(request):
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "vendeur":
        return redirect('vente_home')
    elif person.role == 'gestionnaire':
        return redirect('gestionH')
    logout(request)
    return redirect('login')

@login_required
def vendeur(request):
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    vendeur = Utilisateur.objects.filter(role = 'vendeur')
    if vendeur:
         vendeur = Utilisateur.objects.filter(role = 'vendeur')
         taille = len(vendeur)
         if request.method == "POST":
            mot = request.POST['mot']
            if mot=="":
                 vendeur = Utilisateur.objects.filter(role = 'vendeur')
            else:
                print(mot)
                q_nom = Q(first_name__icontains=mot) if mot else Q()
                q_prenom = Q(last_name__icontains=mot) if mot else Q()
                q_role = Q(role__icontains ='vendeur') if mot else Q()
                q_adr = Q(localites__icontains=mot) if mot else Q()
                q_username = Q(username__icontains=mot) if mot else Q()
                q_sexe = Q(sex=mot) if mot else Q()
                vendeur = Utilisateur.objects.filter((q_nom | q_prenom | q_adr | q_sexe) & q_role) 
                taille = len(vendeur)
         return render(request,'supervisor/vendeur.html',{'vendeurs':vendeur,'taille':taille })
    
    else:
         return render(request,'supervisor/vendeur.html')
    
    
@login_required
def gestionnaire(request):
    # verification de role
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    ges = Utilisateur.objects.filter(role = 'gestionnaire')
    if ges:
         gestion = Utilisateur.objects.filter(role = 'gestionnaire')
         if request.method == "POST":
            mot = request.POST['mot']
            if mot=="":
                gestion = Utilisateur.objects.filter(role = 'gestionnaire')
                print(mot)
            else: 
                q_nom = Q(first_name__icontains=mot) if mot else Q()
                q_prenom = Q(last_name__icontains=mot) if mot else Q()
                q_role = Q(role__icontains ='gestionnaire') if mot else Q()
                q_adr = Q(localites__icontains=mot) if mot else Q()
                q_username = Q(username__icontains=mot) if mot else Q()
                q_sexe = Q(sex=mot) if mot else Q()
                gestion = Utilisateur.objects.filter((q_nom | q_prenom | q_adr | q_sexe) & q_role) 
         taille = len(gestion)
         return render(request,'supervisor/gestionnaire.html',{'gestions':gestion, 'taille': taille}) 
    else:
         return render(request,'supervisor/gestionnaire.html')

@login_required    
def client(request):
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    clients = Client.objects.all()
    taille = len(clients)
    clients = clients[:100]
    if request.method == "POST":
        mot = request.POST['mot']
        q_nom = Q(nom__icontains=mot) if mot else Q()
        q_prenom = Q(prenom__icontains=mot) if mot else Q()
        q_adr = Q(adresse__icontains=mot) if mot else Q()
        q_numero = Q(telephone__icontains=mot) if mot else Q()
        q_sexe = Q(genre=mot) if mot else Q()
        q_matr = Q(matricule=mot) if mot else Q()
        clients = Client.objects.filter(q_nom | q_prenom | q_numero | q_adr | q_sexe | q_matr) 
        taille = len(clients)
    return render(request,'supervisor/client.html', {'clients':clients, 'taille':taille})



@login_required
def desactive_vendeur(request, pk):
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    desactivation = SupervisorUtilisateur.objects.get(id=pk)
    desactivation.is_active = False
    desactivation.save()
    return redirect('liste_vendeur')

@login_required
def active_vendeur(request, pk):
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    desactivation = SupervisorUtilisateur.objects.get(id=pk)
    desactivation.is_active = True
    desactivation.save()
    return redirect('liste_vendeur')

@login_required
def desactive_gestionnaire(request, pk):
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    
    desactivation = SupervisorUtilisateur.objects.get(id=pk)
    desactivation.is_active = False
    desactivation.save()
    return redirect('liste_gestionnaire')

@login_required
def active_gestionnaire(request, pk):
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home ')
    
    desactivation = SupervisorUtilisateur.objects.get(id=pk)
    desactivation.is_active = True
    desactivation.save()
    return redirect('liste_gestionnaire')

@login_required
def dashboard(request):
    # verification de role
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    today = datetime.now()
    client = Client.objects.all()
    produit = Produit.objects.all()
    gestion = Utilisateur.objects.filter(role ="gestionnaire")
    vendeur = Utilisateur.objects.filter(role ="vendeur")
    janvier = 0
    fevrier = 0
    mars =0
    avril=0
    mais=0
    juin=0
    juillet=0
    aout=0
    septembre=0
    octobre=0
    novembre=0
    decembre=0


    janvierb = 0
    fevrierb = 0
    marsb =0
    avrilb=0
    maisb=0
    juinb=0
    juilletb=0
    aoutb=0
    septembreb=0
    octobreb=0
    novembreb=0
    decembreb=0
    nombre_mois = 0
    if today.strftime("%B") == 'January':
        A = [janvier]
        nombre_mois = 1
    elif today.strftime("%B") == 'February' :  
          
          nombre_mois = 2
    elif today.strftime("%B") == 'March':
        
         nombre_mois = 3
    elif today.strftime("%B") == 'April':
               
         nombre_mois = 4
    elif today.strftime("%B") == 'May':
         
         nombre_mois = 5
    elif today.strftime("%B") == 'June':
         
         nombre_mois = 6
    elif today.strftime("%B") == 'July':
        
        nombre_mois = 7
    elif today.strftime("%B") == 'August':
        
        nombre_mois = 8
    elif today.strftime("%B") == 'September':
        
        nombre_mois = 9
    elif today.strftime("%B") == 'October':
       
        nombre_mois = 10
    elif today.strftime("%B") == 'November':
        
        nombre_mois = 11
    elif today.strftime("%B") == 'December': 
        nombre_mois = 12
    
    A = [janvier,fevrier,mars,avril,mais,juin,juillet,aout,septembre,octobre,novembre,decembre]
    B = [janvierb,fevrierb,marsb,avrilb,maisb,juinb,juilletb,aoutb,septembreb,octobreb,novembreb,decembreb]
    P_mois_prec_analyse = 0
    compte = 1
    
    for i in range(nombre_mois+1):
        if i == nombre_mois-1 :
            reste = 13 - i
            l = i
            while l <= reste :
                print(l)
                A[l] = 0
                l+=1   
        P_mois_prec_analyse = VentesAnalyse.objects.annotate(count_ventes=Sum('prix')).values('count_ventes').filter(date__year=today.year,
            date__month = compte)
            
        lol = 0
        for item in P_mois_prec_analyse:  
            lol  = lol + item['count_ventes']    
        A[i] = lol  
        print(A[i])
        compte += 1 
    compte = 0   
    mois = 12
    for i in range(mois): 
        P_mois_prec_analyse = VentesAnalyse.objects.annotate(count_ventes=Sum('prix')).values('count_ventes').filter(date__year=today.year -1,
            date__month = compte)        
        lol = 0
        for item in P_mois_prec_analyse:  
            lol  = lol + item['count_ventes']    
        B[i] = lol  
        print(B[i])
        compte += 1 



    
    vente_annuler = Vente.objects.all().filter(annuler=True)
    vente_annuler = len(vente_annuler)

   
    Profit_journalier = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year,
        date_vente__month=today.month,
        date_vente__day=today.day,
        annuler = False)
    
    
    Profit_journalier_prec = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year,
        date_vente__month=today.month,
        date_vente__day=today.day-1,
        annuler = False)
    
    Profit_journalier_prec2 = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year,
        date_vente__month=today.month,
        date_vente__day=today.day-2,
        annuler = False)
    
    Profit_journalier_prec3 = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year,
        date_vente__month=today.month,
        date_vente__day=today.day-3,
        annuler = False)
    
    Profit_journalier_prec4 = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year,
        date_vente__month=today.month,
        date_vente__day=today.day-4,
        annuler = False
        )
    
    Profit_journalier_prec5 = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year,
        date_vente__month=today.month,
        date_vente__day=today.day-5,
        annuler = False)
    
    total_jour = 0
    for item in Profit_journalier:
        total_jour += item['count_ventes']

    total_jour_prec = 0
    for item in Profit_journalier_prec:
        total_jour_prec += item['count_ventes']

    total_jour_prec2 = 0
    for item in Profit_journalier_prec2:
        total_jour_prec2 += item['count_ventes']

    total_jour_prec3 = 0
    for item in Profit_journalier_prec3:
        total_jour_prec3 += item['count_ventes']

    total_jour_prec4 = 0
    for item in Profit_journalier_prec4:
        total_jour_prec4 += item['count_ventes']

    total_jour_prec5 = 0
    for item in Profit_journalier_prec5:
        total_jour_prec5 += item['count_ventes']
    
    Profit_mesuel = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year,
        date_vente__month=today.month,
        annuler = False)
    
    total_mois = 0
    for item in Profit_mesuel:
        total_mois += item['count_ventes']

    Profit_mesuel_prec = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year,
        date_vente__month=today.month - 1,
        annuler = False )
    
    total_mois_prec = 0
    for item in Profit_mesuel_prec:
        total_mois_prec += item['count_ventes']

    an = today.year

    Profit_anuel = Vente.objects.annotate(count_ventes=Sum('prix_total')).values('count_ventes').filter(date_vente__year=today.year, annuler = False)
    
    total_anuel = 0
    for item in Profit_anuel:
        total_anuel += item['count_ventes']

    Top_client = VentesAnalyse.objects.values('nom').annotate(total=Count('nom')).order_by('-total')[:5]

    top_5_produit =  VentesAnalyse.objects.values('produit').annotate(total=Count('produit')).order_by('-total')[:5]

    # top_5_vendeur = Vente.objects.prefetch_related('id_produit').values('id_produit').annotate(
    # sale_count=Count('id_produit'),
    # produit_nom=Subquery(
    # Produit.objects.filter(id_produit=OuterRef('id_produit')).values('nom_produit')))[:5]


    

    

    # if len(str(total_jour)) == 5:
    #     total_jour = str(total_jour)[:2]+ "K"

    # if len(str(total_mois))== 5:
    #     total_mois = str(total_mois)[:2]+ "K"

    # if len(str(total_jour))>= 6 & len(str(total_jour))<=8:
    #     total_jour = str(total_jour)[:2]+ "M"

    # if len(str(total_mois))>= 6 & len(str(total_mois))>= 8:
    #     total_mois = str(total_mois)[:1]+ "M"
    lundi =0
    mardi=0
    mercredi=0
    jeudi=0
    vendredi = 0
    samedi = 0

    if today.strftime("%A") == 'Monday':
        lundi = total_jour
        mardi = 0
        mercredi = 0
        jeudi = 0
        vendredi = 0
        samedi =0
    
    if today.strftime("%A") == 'Tuesday':
        lundi = total_jour_prec
        mardi = total_jour
        mercredi = 0
        jeudi = 0
        vendredi = 0
        samedi =0 

    if today.strftime("%A") == 'Wednesday':
        lundi = total_jour_prec2 
        mardi = total_jour_prec
        mercredi = total_jour
        jeudi = 0
        vendredi = 0
        samedi =0 
    if today.strftime("%A") == 'Thirsday':
        lundi = total_jour_prec3
        mardi = total_jour_prec2 
        mercredi = total_jour_prec
        jeudi = total_jour
        vendredi = 0
        samedi =0 
    
    if today.strftime("%A") == 'Friday':
        lundi = total_jour_prec4
        mardi = total_jour_prec3
        mercredi = total_jour_prec2
        jeudi = total_jour_prec
        vendredi = total_jour
        samedi =0 
    
    if today.strftime("%A") == 'Saturday':
        lundi = total_jour_prec5
        mardi = total_jour_prec4
        mercredi = total_jour_prec3
        jeudi = total_jour_prec2
        vendredi = total_jour_prec
        samedi =total_jour

    Profit_anuel2021 = VentesAnalyse.objects.annotate(count_ventes=Sum('prix')).values('count_ventes').filter(date__year=2021)
    total_anuel2021 = 0
    for item in Profit_anuel2021:
        total_anuel2021 += item['count_ventes']

    Profit_anuel2022 = VentesAnalyse.objects.annotate(count_ventes=Sum('prix')).values('count_ventes').filter(date__year=2022)
    total_anuel2022 = 0
    for item in Profit_anuel2022:
        total_anuel2022 += item['count_ventes']
    
    Profit_anuel2023 = VentesAnalyse.objects.annotate(count_ventes=Sum('prix')).values('count_ventes').filter(date__year=2023)
    total_anuel2023 = 0
    for item in Profit_anuel2023:
        total_anuel2023 += item['count_ventes']

    Profit_anuel2024 = VentesAnalyse.objects.annotate(count_ventes=Sum('prix')).values('count_ventes').filter(date__year=2024)
    total_anuel2024 = 0
    for item in Profit_anuel2024:
        total_anuel2024 += item['count_ventes']
    
    
    profit_anuel_prec = VentesAnalyse.objects.annotate(count_ventes=Sum('prix')).values('count_ventes').filter(date__year=(today.year -1))
    
       
    total_anuel_prec = 0
    for item in profit_anuel_prec:
        total_anuel_prec += item['count_ventes']

    if total_anuel_prec == 0:
        croissance = 0
    else:
        croissance = ((total_anuel2024 - total_anuel_prec)/total_anuel_prec)*100

    
    
    if total_jour_prec == 0:
        croissance_jour = 0
    else:
        croissance_jour = ((total_jour - total_jour_prec)/total_jour_prec)*100

    if total_jour_prec == 0:
        croissance_mois = 0
    else:
        croissance_mois = ((total_mois - total_mois_prec)/total_mois_prec)*100
    
    print(today.strftime('%B'))
    heure = today.hour



    return render(request, 'supervisor/dashboard.html', {'client':len(client),
                                                         'produit':len(produit),
                                                         'gestionnaire': len(gestion),
                                                         'vendeur':len(vendeur),
                                                         'total_jour': total_jour,
                                                         'total_mois':total_mois,
                                                         'an':str(an),
                                                         'total_anuel':total_anuel,
                                                         'top_client':Top_client,
                                                         'top_produit_5': top_5_produit,
                                                         'total_jour_prec':int(total_jour_prec),
                                                         'lundi':lundi,
                                                        'mardi':mardi,
                                                        'mercredi':mercredi,
                                                        'jeudi':jeudi,
                                                        'vendredi':vendredi,
                                                        'samedi':samedi,
                                                        "total_anuel2021":total_anuel2021,
                                                        "total_anuel2022":total_anuel2022,
                                                        "total_anuel2023":total_anuel2023,
                                                        "total_anuel2024":total_anuel2024,
                                                        "croissance":int(croissance),
                                                        "croissance_jour":int(croissance_jour),
                                                        "croissance_mois":int(croissance_mois),
                                                        "vente_annuler":vente_annuler,
                                                        "janvier":A[0],
                                                        "fev":A[1],
                                                        "mars":A[2],
                                                        "avril":A[3],
                                                        "mais":A[4],
                                                        "juin":A[5],
                                                        "juillet":A[6],
                                                        "aout":A[7],
                                                        "sep":A[8],
                                                        "octobre":A[9],
                                                        "nov":A[10],
                                                        "dec":A[11],

                                                        "janvierB":B[0],
                                                        "fevB":B[1],
                                                        "marsB":B[2],
                                                        "avrilB":B[3],
                                                        "maisB":B[4],
                                                        "juinB":B[5],
                                                        "juilletB":B[6],
                                                        "aoutB":B[7],
                                                        "sepB":B[8],
                                                        "octobreB":B[9],
                                                        "novB":B[10],
                                                        "decB":B[11],
                                                        "heure":heure,
                                                        
                                                        })

def prediction_deco(request):
    ventes_prediction = VentesAnalyse.objects.values('date', 'quantite', 'produit')

    # Convertir les données en DataFrame pandas
    df = pd.DataFrame(ventes_prediction)

    # Convertir la colonne de date en datetime
    df['date'] = pd.to_datetime(df['date'])

    # Renommer les colonnes pour correspondre à ce que Prophet attend
    df = df.rename(columns={'date': 'ds', 'quantite': 'y', 'produit': 'product'})

    # Supprimer les informations de fuseau horaire de la colonne 'ds'
    df['ds'] = df['ds'].dt.tz_localize(None)

    # Créer une liste des produits uniques
    products = df['product'].unique()

    # Dictionnaire pour stocker les prédictions de chaque produit
    predictions = {}
    best_selling_product = None

    # Pour chaque produit, entraîner un modèle et prédire les ventes en avril 2024
    for product in products:
        # Filtrer les données pour le produit actuel
        product_df = df[df['product'] == product]

        # Créer un modèle Prophet
        model = Prophet()

        # Entraîner le modèle sur les données historiques du produit
        model.fit(product_df[['ds', 'y']])

        # Créer un DataFrame pour les dates de prédiction (avril 2024)
        future_dates = pd.DataFrame({'ds': pd.date_range(start='2024-04-01', end='2024-04-30')})

        # Faire la prédiction
        forecast = model.predict(future_dates)

        # Calculer le total des quantités prévues pour le produit en avril 2024
        total_predicted_sales = forecast['yhat'].sum()

        # Stocker la prédiction dans le dictionnaire
        predictions[product] = total_predicted_sales

    # Trouver le produit avec les ventes prévues les plus élevées
    best_selling_product = max(predictions, key=predictions.get)
    return render(request,"supervisor/prediction.html",{'best_selling_product': best_selling_product,
                                                    'predicted_sales': int(predictions[best_selling_product])})


def test(request):
    best_selling_product = None
    lo = 0

    return render(request,"supervisor/test.html",{'best_selling_product': best_selling_product,
                                                    'lol':lo})


# Create your views here.




# def remplisage(request):
#     STATIC_DIR = os.path.join(settings.BASE_DIR, 'data')  # Assuming BASE_DIR is defined

# # Get the full path to the CSV file
#     csv_file_path = os.path.join(STATIC_DIR, 'donnee_vente.csv')

# # Open the CSV file in read mode
#     with open(csv_file_path, 'r') as csvfile:
#         reader = csv.DictReader(csvfile)

#     # Iterate through each row in the CSV file
#         for row in reader:
#             model_instance = VentesAnalyse(
#             nom =row['Name'],
#             genre = row['Gender'],  # Replace with actual field names
#             produit = row['Produit'],
#             type = row['Type'],
#             prix = row['Prix'],
#             date = row['Date']
#                 # ... and so on for each fiel
#                 )
#             model_instance.save()
   
  
#     return(request,"supervisor/try.html" )

@login_required
def export_data_client_to_csv(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database (replace with your actual data retrieval logic)
    data = Client.objects.all()  # Assuming this function returns a list of data rows

    # Create a CSV file object in memory
    csv_file = csv.writer(open('Client.csv', 'w', newline=''))

    # Write header row
    csv_file.writerow(['Id','Nom', 'Prenom', 'Adresse','Telephone','Email','Sexe','Matricule'])  # Replace with actual header names
    print(data)
    # Write data rows
    for d in data:     
        csv_file.writerow([d.id_client,d.nom,d.prenom, d.adresse, d.telephone,d.email,d.genre,d.matricule])

    # Prepare the Django response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="client.csv"'

    # Write CSV content to the response
    writer = csv.writer(response)
    writer.writerow(['Id','Nom', 'Prenom', 'Adresse','Telephone','Email','Sexe','Matricule'])  # Replace with actual header names
    for d in data:
        writer.writerow([d.id_client,d.nom,d.prenom, d.adresse, d.telephone,d.email,d.genre,d.matricule])

    return response



@login_required
def export_data_client_to_excel(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database
    data = Client.objects.all()  # Assuming this is a QuerySet of Produit models

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active  # Get the active worksheet

    # Write header row
    header = ['Id','Nom', 'Prenom', 'Adresse','Telephone','Email','Sexe','Matricule']
    worksheet.append(header)

    # Write data rows
    for d in data:
        row_data = [d.id_client,d.nom,d.prenom, d.adresse, d.telephone,d.email,d.genre,d.matricule]
        worksheet.append(row_data)

    # Prepare the Django response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Client.xlsx"'

    # Write Excel data to the response
    workbook.save(response)

    return response


@login_required
def export_data_vendeur_to_csv(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database (replace with your actual data retrieval logic)
    data = SupervisorUtilisateur.objects.filter(role = "vendeur")  # Assuming this function returns a list of data rows

    # Create a CSV file object in memory
    csv_file = csv.writer(open('vendeur.csv', 'w', newline=''))

    # Write header row
    csv_file.writerow(['Id','Nom', 'Prenom', 'Adresse','Email','Sexe','Telephone','Role'])  # Replace with actual header names
    print(data)
    # Write data rows
    for d in data:     
        csv_file.writerow([d.id,d.first_name,d.last_name, d.localites, d.email,d.sex,d.tel,d.role])

    # Prepare the Django response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="vendeur.csv"'

    # Write CSV content to the response
    writer = csv.writer(response)
    writer.writerow(['Id','Nom', 'Prenom', 'Adresse','Email','Sexe','Telephone','Role'])  # Replace with actual header names
    for d in data:
        writer.writerow([d.id,d.first_name,d.last_name, d.localites, d.email,d.sex,d.tel,d.role])

    return response

@login_required
def export_data_vendeur_to_excel(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database
    data = SupervisorUtilisateur.objects.filter(role = "vendeur")  # Assuming this is a QuerySet of Produit models

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active  # Get the active worksheet

    # Write header row
    header = ['Id','Nom', 'Prenom', 'Adresse','Email','Sexe','Telephone','Role']
    worksheet.append(header)

    # Write data rows
    for d in data:
        row_data = [d.id,d.first_name,d.last_name, d.localites, d.email,d.sex,d.tel,d.role]
        worksheet.append(row_data)

    # Prepare the Django response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="vendeur.xlsx"'

    # Write Excel data to the response
    workbook.save(response)

    return response


@login_required
def export_data_gestion_to_csv(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database (replace with your actual data retrieval logic)
    data = SupervisorUtilisateur.objects.filter(role = "gestionnaire")  # Assuming this function returns a list of data rows

    # Create a CSV file object in memory
    csv_file = csv.writer(open('gestionnaire.csv', 'w', newline=''))

    # Write header row
    csv_file.writerow(['Id','Nom', 'Prenom', 'Adresse','Email','Sexe','Telephone','Role'])  # Replace with actual header names
    print(data)
    # Write data rows
    for d in data:     
        csv_file.writerow([d.id,d.first_name,d.last_name, d.localites, d.email,d.sex,d.tel,d.role])

    # Prepare the Django response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="gestionnaire.csv"'

    # Write CSV content to the response
    writer = csv.writer(response)
    writer.writerow(['Id','Nom', 'Prenom', 'Adresse','Email','Sexe','Telephone','Role'])  # Replace with actual header names
    for d in data:
        writer.writerow([d.id,d.first_name,d.last_name, d.localites, d.email,d.sex,d.tel,d.role])

    return response

@login_required
def export_data_gestion_to_excel(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database
    data = SupervisorUtilisateur.objects.filter(role = "gestionnaire")  # Assuming this is a QuerySet of Produit models

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active  # Get the active worksheet

    # Write header row
    header = ['Id','Nom', 'Prenom', 'Adresse','Email','Sexe','Telephone','Role']
    worksheet.append(header)

    # Write data rows
    for d in data:
        row_data = [d.id,d.first_name,d.last_name, d.localites, d.email,d.sex,d.tel,d.role]
        worksheet.append(row_data)

    # Prepare the Django response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="gestionnaire.xlsx"'

    # Write Excel data to the response
    workbook.save(response)

    return response



@login_required
def export_data_vente_to_csv(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database (replace with your actual data retrieval logic)
    data = VentesAnalyse.objects.all() # Assuming this function returns a list of data rows

    # Create a CSV file object in memory
    csv_file = csv.writer(open('ventes.csv', 'w', newline=''))

    # Write header row
    csv_file.writerow(['Nom', 'Genre', 'Produit','Categorie','Prix','Date'])  # Replace with actual header names
    print(data)
    # Write data rows
    for d in data:     
        csv_file.writerow([d.nom,d.genre, d.produit, d.type,d.prix,d.date])

    # Prepare the Django response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ventes.csv"'

    # Write CSV content to the response
    writer = csv.writer(response)
    writer.writerow(['Nom', 'Genre', 'Produit','Categorie','Prix','Date'])  # Replace with actual header names
    for d in data:
        writer.writerow([d.nom,d.genre, d.produit, d.type,d.prix,d.date])

    return response

@login_required
def export_data_vente_to_excel(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database
    data = VentesAnalyse.objects.all() # Assuming this is a QuerySet of Produit models

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active  # Get the active worksheet

    # Write header row
    header = ['Nom', 'Genre', 'Produit','Categorie','Prix','Date']
    worksheet.append(header)

    # Write data rows
    for d in data:
        row_data = [d.nom,d.genre, d.produit, d.type,d.prix,str(d.date)]
        worksheet.append(row_data)

    # Prepare the Django response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ventes.xlsx"'

    # Write Excel data to the response
    workbook.save(response)

    return response




@login_required
def export_data_vente_annuler_to_excel(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database
    data = VenteAnuler.objects.prefetch_related('id_vente') # Assuming this is a QuerySet of Produit models

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active  # Get the active worksheet

    # Write header row
    header = ['n°','Nom du client', 'Produit','Categorie','Quantité','Prix','Nom du vendeur','Date','Motif_annulation']
    worksheet.append(header)

    # Write data rows
    for d in data:
        row_data = [d.id_vente_anuler,(d.id_vente.id_client.nom + " " + d.id_vente.id_client.prenom),d.id_vente.id_produit.nom_produit, d.id_vente.id_produit.categorie, d.id_vente.quantite_vendu,d.id_vente.prix_total,(d.id_vente.id_vendeur.first_name + " " + d.id_vente.id_vendeur.last_name),str(d.id_vente.date_vente),d.motif_annulation]
        worksheet.append(row_data)

    # Prepare the Django response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ventes_annuler.xlsx"'

    # Write Excel data to the response
    workbook.save(response)

    return response


@login_required
def export_data_vente_to_excel(request):
    person = SupervisorUtilisateur.objects.get(id=request.user.id)
    if person.role == "gestionnaire":
        return redirect('gestionH')
    elif person.role == 'vendeur':
        return redirect('vente_home')
    # Retrieve data from the database
    data = VentesAnalyse.objects.all() # Assuming this is a QuerySet of Produit models

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active  # Get the active worksheet

    # Write header row
    header = ['Nom', 'Genre', 'Produit','Categorie','Prix','Date']
    worksheet.append(header)

    # Write data rows
    for d in data:
        row_data = [d.nom,d.genre, d.produit, d.type,d.prix,str(d.date)]
        worksheet.append(row_data)

    # Prepare the Django response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ventes.xlsx"'

    # Write Excel data to the response
    workbook.save(response)

    return response


def export_data_produit_to_excel(request):
    # Retrieve data from the database
    data = Produit.objects.all()  # Assuming this is a QuerySet of Produit models

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active  # Get the active worksheet

    # Write header row
    header = ['id', 'produit', 'categorie', 'prix', 'quantite']
    worksheet.append(header)

    # Write data rows
    for product in data:
        row_data = [product.id_produit, product.nom_produit, product.categorie, product.p_unitaire, product.quantite]
        worksheet.append(row_data)

    # Prepare the Django response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="produit.xlsx"'

    # Write Excel data to the response
    workbook.save(response)

    return response